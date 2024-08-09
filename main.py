""" import pandas as pd

# Cargar el archivo Excel
archivo_excel = 'C:/Equipo/Curso/Jorge/Manejo de datos/1.xlsx'
df = pd.read_excel(archivo_excel)
print(df.head())



casillas_marcadas = df[df['si'] == True]
print(casillas_marcadas) """


from openpyxl import load_workbook


# Cargar el libro de trabajo y seleccionar la hoja activa
wb = load_workbook('C:/Equipo/Curso/Jorge/Manejo de datos/1.xlsx')
ws = wb.active
print(ws)

# Supongamos que sabes las celdas específicas donde están los valores de las casillas de verificación
# Por ejemplo, si están en la columna A desde la fila 1 hasta la 10
""" for row in range(1, 11):
    cell_value = ws[f'A{row}'].value
    print(f'Valor en la celda A{row}: {cell_value}') """
nombre =  ws[f'B4'].value
print(nombre)

verticalSi = ws[f'H6'].value
verticalNo = ws[f'H7'].value
print(verticalSi)
print(verticalNo)